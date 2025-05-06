import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.utils.units import cm_to_EMU
from io import BytesIO
from PIL import Image as PILImage
from openpyxl.drawing.xdr import XDRPositiveSize2D
import io
import base64

# --- Configuración de las dimensiones del área de la imagen (en cm) ---
AREA_HEIGHT_CM = 6.8
AREA_WIDTH_CM = 9.42

# ---  Función rotar imagen
def rotate_image(image_bytes, angle):
    """Rota una imagen en bytes y la devuelve en bytes."""
    img = PILImage.open(io.BytesIO(image_bytes))
    rotated_img = img.rotate(angle, expand=True)
    img_buffer = io.BytesIO()
    rotated_img.save(img_buffer, format=img.format)
    return img_buffer.getvalue()

# --- Función para redimensionar la imagen manteniendo la relación de aspecto ---
def redimensionar_imagen(imagen_pil, max_ancho_cm, max_alto_cm, dpi=96):
    """Redimensiona una imagen de Pillow manteniendo su relación de aspecto."""
    max_ancho_pixels = max_ancho_cm * dpi / 2.54
    max_alto_pixels = max_alto_cm * dpi / 2.54
    ancho, alto = imagen_pil.size

    ratio_ancho = max_ancho_pixels / ancho
    ratio_alto = max_alto_pixels / alto

    if ratio_ancho < 1 or ratio_alto < 1:  # Solo redimensionar si es más grande
        ratio = min(ratio_ancho, ratio_alto)
        nuevo_ancho = int(ancho * ratio)
        nuevo_alto = int(alto * ratio)
        imagen_redimensionada = imagen_pil.resize((nuevo_ancho, nuevo_alto))
        return imagen_redimensionada
    return imagen_pil

# --- Función para convertir cm a EMUs ---
c2e = cm_to_EMU

# --- Función para calcular el offset en EMUs para centrar la imagen ---
def calcular_offset(area_cm, img_cm):
    delta_cm = ((area_cm - img_cm) / 2)+0.1
    return c2e(delta_cm)

# --- Interfaz de Usuario en Streamlit ---
st.title("Registro fotografico")

formato = ["Preventivo", "Recorredor", "clientes interno", "clientes externo","Factibilidades"]
formato_seleccionado = st.radio("Selecciona el formato:", formato)

if formato_seleccionado == "Factibilidades":
    AREA_HEIGHT_CM = 5.44
    AREA_WIDTH_CM = 7.76


opciones = ["BRAYAN STIVEN SALAMANCA CASTAÑEDA", "JUAN CARLOS RODRIGUEZ CHIQUILLO", "YOVANNI GOMEZ PEÑA", "MICHAEL ESTEBAN URQUIJO LOPEZ", 
            "FABIAN DAVID CIFUENTES GRUESO","JOHN EZEQUIEL TANGARIFE ARENAS","ELIAS ALBERTO CADENA AGUDELO","NICOLAS EDUARDO ALVARADO CASTAÑO",
            "DEISON RAUL HURTADO MOSQUERA","JHON HAMILTON BARRERA ALARCON","NEIDER ALEXANDER ARENAS MORALES","EDISON ESTEBAN SEPULVEDA VALENCIA",
            "JOHN ALEXANDER VEGA SALAMANCA"
            ]
map_telefono = {
            "BRAYAN STIVEN SALAMANCA CASTAÑEDA":"3133977853", 
            "JUAN CARLOS RODRIGUEZ CHIQUILLO":"3214522373", 
            "YOVANNI GOMEZ PEÑA":"3112973928", 
            "MICHAEL ESTEBAN URQUIJO LOPEZ":"3114669376", 
            "FABIAN DAVID CIFUENTES GRUESO":"3112042566",
            "JOHN EZEQUIEL TANGARIFE ARENAS":"3118859551",
            "ELIAS ALBERTO CADENA AGUDELO":"3102619401",
            "NICOLAS EDUARDO ALVARADO CASTAÑO":"3208216107",
            "DEISON RAUL HURTADO MOSQUERA":"3202366550",
            "JHON HAMILTON BARRERA ALARCON":"3208550790",
            "NEIDER ALEXANDER ARENAS MORALES":"3114890889",
            "EDISON ESTEBAN SEPULVEDA VALENCIA":"3154928823",
            "JOHN ALEXANDER VEGA SALAMANCA":"3219662645"
            }

ejecutor = st.selectbox("Ejecutor:", opciones)

if formato_seleccionado == "clientes interno" or formato_seleccionado == "clientes externo" or formato_seleccionado == "Factibilidades":
    cliente = st.text_input("Nombre del sitio:")

telefono = map_telefono.get(ejecutor, "")
direccion = st.text_input("DIRECCIÓN:")
fecha_visita = st.date_input("FECHA DE LA VISITA:")
telefono = st.text_input("TELEFONO:", value=telefono, disabled=True)

uploaded_files = st.file_uploader("Subir Registros Fotográficos", accept_multiple_files=True, type=["png", "jpg", "jpeg"])
descripciones = [""] * len(uploaded_files)

# --- Sección de Vista Previa con Miniaturas y Campos de Descripción Adyacentes ---
if ejecutor or direccion or fecha_visita or telefono or uploaded_files:
    st.subheader("Vista Previa de los Datos:")
    if ejecutor:
        st.write(f"**Ejecutor:** {ejecutor}")
    if direccion:
        st.write(f"**Dirección:** {direccion}")
    if fecha_visita:
        st.write(f"**Fecha de la Visita:** {fecha_visita.strftime('%Y-%m-%d')}")
    if telefono:
        st.write(f"**Teléfono:** {telefono}")
    if uploaded_files:
        st.write("**Registros Fotográficos:**")
        if formato_seleccionado == "Factibilidades":
            num_filas_preview = (len(uploaded_files) + 2) // 3
            for i in range(num_filas_preview):
                cols = st.columns(3)
                for j in range(3):
                    idx = i * 3 + j
                    if idx < len(uploaded_files):
                        file = uploaded_files[idx]
                        key_rotacion = f"rotacion_{idx}"
                        key_imagen_rotada = f"imagen_rotada_{idx}"

                        if key_rotacion not in st.session_state:
                            st.session_state[key_rotacion] = 0

                        col_imagen_botones = st.columns([3, 2])
                        with col_imagen_botones[0]:
                            try:
                                img = PILImage.open(file)
                                if st.session_state[key_rotacion] != 0:
                                    rotated_img = img.rotate(st.session_state[key_rotacion], expand=True)
                                    st.image(rotated_img, caption=f"Foto {idx+1} (Rotada {st.session_state[key_rotacion]}°)", width=100)
                                else:
                                    st.image(img, caption=f"Foto {idx+1}", width=100)
                            except Exception as e:
                                st.error(f"Error: No se pudo abrir el archivo como imagen: {file.name}")

                        with col_imagen_botones[1]:
                            col_rot_left, col_rot_right = st.columns(2)
                            with col_rot_left:
                                if st.button("↺", key=f"rotar_der_{idx}"):
                                    st.session_state[key_rotacion] = (st.session_state[key_rotacion] + 90) % 360
                                    st.rerun()                                
                            with col_rot_right:
                                if st.button("↻", key=f"rotar_izq_{idx}"):
                                    st.session_state[key_rotacion] = (st.session_state[key_rotacion] - 90) % 360
                                    st.rerun()
                                
                descripcion_key = f"descripcion_factibilidad_{i}"
                descripciones.extend([""] * 3)
                descripciones[i * 3: (i + 1) * 3] = [st.text_input(f"Descripción para Fotos {(i * 3) + 1} a {(i + 1) * 3}:", key=descripcion_key)] * 3
        else:
            for i, file in enumerate(uploaded_files):
                key_rotacion = f"rotacion_{i}"
                key_imagen_rotada = f"imagen_rotada_{i}"

                if key_rotacion not in st.session_state:
                    st.session_state[key_rotacion] = 0

                col_imagen_botones = st.columns([3, 1])
                with col_imagen_botones[0]:
                    try:
                        img = PILImage.open(file)
                        if st.session_state[key_rotacion] != 0:
                            rotated_img = img.rotate(st.session_state[key_rotacion], expand=True)
                            st.image(rotated_img, caption=f"Foto {i+1} (Rotada {st.session_state[key_rotacion]}°)", width=100)
                        else:
                            st.image(img, caption=f"Foto {i+1}", width=100)
                    except Exception as e:
                        st.error(f"Error: No se pudo abrir el archivo como imagen: {file.name}")

                with col_imagen_botones[1]:
                    col_rot_left, col_rot_right = st.columns(2)
                    with col_rot_left:
                        if st.button("↺", key=f"rotar_der_{i}"):
                            st.session_state[key_rotacion] = (st.session_state[key_rotacion] + 90) % 360
                            st.rerun()
                        
                    with col_rot_right:
                        if st.button("↻", key=f"rotar_izq_{i}"):
                            st.session_state[key_rotacion] = (st.session_state[key_rotacion] - 90) % 360
                            st.rerun()
                        

                descripciones[i] = st.text_input(f"Descripción para la Foto {i+1}:", key=f"descripcion_{i}")



if st.button("Generar Excel"):
    if ejecutor and direccion and fecha_visita and telefono and uploaded_files:
        try:
            ruta_excel = ""
            fila_foto_inicio = 8
            columna_foto_inicio = 1  # Columna A
            celda_ejecutor = 'G7'
            celda_dirección = 'C5'
            celda_fecha = 'C6'
            celda_tel = 'H7'

            if formato_seleccionado == "Preventivo" or formato_seleccionado == "Recorredor":
                ruta_excel = 'RF_PREVENTIVO.XLSX'

            elif formato_seleccionado == "clientes interno":
                ruta_excel = 'RF_CLIENTE_INTERNO.xlsx'
                fila_foto_inicio = 10
                celda_ejecutor = 'G8'
                celda_dirección = 'C6'
                celda_fecha = 'C7'
                celda_tel = 'H8'

            elif formato_seleccionado == "clientes externo":
                ruta_excel = 'RF_CLIENTE_EXTERNO.xlsx'
                fila_foto_inicio = 10
                celda_ejecutor = 'G8'
                celda_dirección = 'C6'
                celda_fecha = 'C7'
                celda_tel = 'H8'

            elif formato_seleccionado == "Factibilidades":
                ruta_excel = 'RF_FACTIBILIDADES.xlsx'
                fila_foto_inicio = 12
                columna_foto_inicio = 1 # Inicial en A
                celda_ejecutor = 'B9' # Ajusta según tu formato
                celda_dirección = 'B7' # Ajusta según tu formato
                celda_fecha = 'B8'   # Ajusta según tu formato
                celda_tel = 'D9'     # Ajusta según tu formato

            libro = load_workbook(ruta_excel)
            hoja = libro.active

            # --- Llenar los campos de texto ---
            if formato_seleccionado == "Recorredor":
                hoja['A4'] = 'REGISTRO FOTOGRÁFICO RECORREDOR'
                hoja['D7'] = 'RECORREDOR'

            hoja[celda_ejecutor] = ejecutor
            hoja[celda_dirección] = direccion
            hoja[celda_fecha] = fecha_visita.strftime("%d-%m-%Y")
            hoja[celda_tel] = telefono
            if formato_seleccionado == "clientes interno" or formato_seleccionado == "clientes externo":
                hoja['C5'] = cliente
            if formato_seleccionado == "Factibilidades":
                hoja['B6'] = cliente

            fila_actual_foto = fila_foto_inicio
            for i, archivo_subido in enumerate(uploaded_files):
                key_rotacion = f"rotacion_{i}"
                angulo_rotacion = st.session_state.get(key_rotacion, 0) # Obtener el ángulo de rotación

                 # --- Rotar la imagen antes de redimensionar ---
                img_pil_original = PILImage.open(archivo_subido)
                img_pil_rotada = img_pil_original.rotate(angulo_rotacion, expand=True)


                # --- Redimensionar la imagen ---
                #img_pil = PILImage.open(archivo_subido)
                img_redimensionada = redimensionar_imagen(img_pil_rotada, AREA_WIDTH_CM, AREA_HEIGHT_CM)
                img_width_cm, img_height_cm = img_redimensionada.size[0] * 2.54 / 96, img_redimensionada.size[1] * 2.54 / 96

                # --- Convertir la imagen para openpyxl ---
                img_buffer = BytesIO()
                img_redimensionada.save(img_buffer, format="PNG")
                img_buffer.seek(0)
                img = Image(img_buffer)

                # --- Calcular la columna de anclaje para Factibilidades ---
                if formato_seleccionado == "Factibilidades":
                    if i % 3 == 0:
                        col_idx = 0  # Columna A
                    elif i % 3 == 1:
                        col_idx = 3  # Columna D
                    else:
                        col_idx = 6  # Columna G
                else: # Para otros formatos, la lógica original
                    col_idx = (columna_foto_inicio - 1) + (4 * (i % 2))

                row_idx = fila_actual_foto - 1

                x_offset_emu = calcular_offset(AREA_WIDTH_CM, img_width_cm)
                y_offset_emu = calcular_offset(AREA_HEIGHT_CM, img_height_cm)

                # --- Definir el marcador de anclaje ---
                marker = AnchorMarker(col=col_idx, colOff=x_offset_emu, row=row_idx, rowOff=y_offset_emu)
                size = XDRPositiveSize2D(cx=c2e(img_width_cm), cy=c2e(img_height_cm))
                img.anchor = OneCellAnchor(_from=marker, ext=size)
                hoja.add_image(img)

                # --- Calcular la celda unificada para la descripción para Factibilidades ---
                if formato_seleccionado == "Factibilidades":
                    celda_descripcion_inicio = f"B{fila_actual_foto + 1}"
                    celda_descripcion_fin = f"H{fila_actual_foto + 2}"
                    hoja.merge_cells(f"{celda_descripcion_inicio}:{celda_descripcion_fin}")
                    hoja[celda_descripcion_inicio] = descripciones[i]
                    if (i + 1) % 3 == 0: # Saltar cada 3 imágenes (al final de la fila A, D, G)
                        fila_actual_foto += 6
                else: # Lógica de descripción para otros formatos
                    if (i + 1) % 2 != 0:
                        celda_descripcion_inicio = f"B{fila_actual_foto + 15}"
                        celda_descripcion_fin = f"D{fila_actual_foto + 16}"
                    else:
                        celda_descripcion_inicio = f"F{fila_actual_foto + 15}"
                        celda_descripcion_fin = f"H{fila_actual_foto + 16}"
                        fila_actual_foto += 17
                    hoja.merge_cells(f"{celda_descripcion_inicio}:{celda_descripcion_fin}")
                    hoja[celda_descripcion_inicio] = descripciones[i]

            # --- Guardar y ofrecer descarga ---
            buffer = BytesIO()
            libro.save(buffer)
            buffer.seek(0)
            st.download_button(
                label="Descargar Excel Generado",
                data=buffer,
                file_name=f"Registro_fotografico_{fecha_visita.strftime('%d-%m-%Y')} {direccion}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            st.success("¡El archivo Excel ha sido generado exitosamente!")

        except Exception as e:
            st.error(f"Ocurrió un error: {e}")
    else:
        st.warning("Por favor, completa todos los campos y sube al menos una foto.")
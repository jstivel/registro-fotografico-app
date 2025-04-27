import streamlit as st
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from io import BytesIO

# --- Interfaz de Usuario en Streamlit ---
st.title("Aplicación para Llenar Formato Excel")

ejecutor = st.text_input("EJECUTOR:")
direccion = st.text_input("DIRECCIÓN:")
fecha_visita = st.date_input("FECHA DE LA VISITA:")
telefono = st.text_input("TELEFONO:")

uploaded_files = st.file_uploader("Subir Registros Fotográficos", accept_multiple_files=True, type=["png", "jpg", "jpeg"])
descripciones = []
for i, file in enumerate(uploaded_files):
    descripcion = st.text_input(f"Descripción para la Foto {i+1}:")
    descripciones.append(descripcion)

if st.button("Generar Excel"):
    if ejecutor and direccion and fecha_visita and telefono and uploaded_files:
        try:
            # --- Cargar el archivo Excel existente ---
            ruta_excel = 'REGISTRO FOTOGRAFICO.XLSX'  # Reemplaza con la ruta de tu archivo
            libro = load_workbook(ruta_excel)
            hoja = libro.active

            # --- Llenar los campos de texto ---
            hoja['G7'] = ejecutor
            hoja['C5'] = direccion
            hoja['C6'] = fecha_visita.strftime("%Y-%m-%d") # Formatear la fecha
            hoja['H7'] = telefono

            # --- Insertar imágenes y descripciones ---
            fila_foto = 8
            columna_foto = 1  # Inicia en la columna A
            fila_descripcion = 23
            columna_descripcion = 2 # Inicia en la columna B

            for i, archivo_subido in enumerate(uploaded_files):
                img = Image(BytesIO(archivo_subido.read()))

                # Calcular la celda para la foto
                celda_foto = f"{chr(ord('A') + (columna_foto - 1))}{fila_foto}"
                hoja.add_image(img, celda_foto)

                # Calcular la celda unificada para la descripción
                celda_descripcion_inicio = f"{chr(ord('B') + (columna_descripcion - 2))}{fila_descripcion}"
                celda_descripcion_fin = f"{chr(ord('D') + (columna_descripcion - 2))}{fila_descripcion + 1}"
                hoja.merge_cells(f"{celda_descripcion_inicio}:{celda_descripcion_fin}")
                hoja[celda_descripcion_inicio] = descripciones[i]

                # Actualizar las filas y columnas para la siguiente foto
                if (i + 1) % 2 == 0: # Después de cada dos fotos
                    fila_foto += 17
                    fila_descripcion += 17
                    columna_foto = 1 # Reinicia a la columna A
                    columna_descripcion = 2 # Reinicia a la columna B
                else:
                    columna_foto = 5 # Salta a la columna E
                    columna_descripcion = 6 # Salta a la columna F

            # --- Guardar el archivo modificado en memoria ---
            buffer = BytesIO()
            libro.save(buffer)
            buffer.seek(0)

            # --- Ofrecer la descarga ---
            st.download_button(
                label="Descargar Excel Generado",
                data=buffer,
                file_name="formato_llenado.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.success("¡El archivo Excel ha sido generado exitosamente!")

        except Exception as e:
            st.error(f"Ocurrió un error: {e}")
    else:
        st.warning("Por favor, completa todos los campos y sube al menos una foto.")